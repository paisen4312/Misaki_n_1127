import tweepy
import csv
import participants
import datetime
import openpyxl
import xlwt
import pandas as pd
import properties
import urllib
from bs4 import BeautifulSoup
import bs4
import requests


#フロントから叩くようの設定
class ope:
    # 各種ツイッターのキーをセット
    consumer_key='pyyAx9RMsE4OheRf8MOYPe2pA'
    consumer_secret='Ai39VmEhftFQDhcCRitA8zBS1beheqMOyHfua1saXHKQl4NbER'
    access_token_key='1292101677101678593-Gmso0N0qBAj6WeUPtjSbMqILqoDp1m'
    access_token_secret='01btngKQSNCx6cDhjAltFXkESgbZlY5zB8b9ciyWw8VSP'

    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_token_key, access_token_secret)
    api = tweepy.API(auth)

    summary_data = []

    def makeSheets():

        wb = openpyxl.Workbook(properties.excel_path)

        participants.makeElemsDict()
        wb.create_sheet("集計シート")
        wb.create_sheet("配信")
        wb.create_sheet("RT")

        for user in participants.elems_dict.values():
            print(user)
            wb.create_sheet(user)

        wb.save(properties.excel_path)

    #リストを持っているユーザーのcsvを作成する
    def hasListUser():
        #ユーザー情報dict作成
        participants.makeListUsers()
        RT_count = 0
        #1人ずつ処理するためのfor文
        for user in participants.dict.keys():
            #集計開始ログ出力
            dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')
            print("【"+ dt_now_str +"】"+ participants.dict[user] + "の集計開始")
            #ツイート情報格納用の箱
            tweet_data = []
            #ツイートごとに格納するためのやつ
            for tweet in tweepy.Cursor(api.user_timeline, screen_name = user ,exclude_replies = True).items():
                #時間を日本時間に設定
                tweetDayTime = tweet.created_at + datetime.timedelta(hours=9)
                #時間をstr型に変換
                tdt_str=tweetDayTime.strftime('%Y/%m/%d_%H:%M:%S')
                #イベント開始時よりツイート日時が新しい場合、ツイート情報を箱に追加
                if tweet.created_at > properties.start_date:
                    tweet_data.append([tdt_str,tweet.text.replace('\n',''),tweet.retweet_count])
                    RT_count = RT_count + tweet.retweet_count
            #サマリシートに集計数を追加
            summary_data.append([participants.dict[user], RT_count])
            #RT数をリセット
            RT_count = 0

            #csv出力
            with open(properties.csv_path + participants.dict[user] +'.csv', 'w',newline='',encoding='utf-8') as f:
                writer = csv.writer(f, lineterminator='\n')
                writer.writerow(["created_at","text","RT"])
                writer.writerows(tweet_data)
                dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')
                print("【"+ dt_now_str +"】"+ participants.dict[user] + "の集計完了")
            
            with open(properties.csv_path + 'summary.csv', 'w',newline='',encoding='utf-8') as f:
                writer = csv.writer(f, lineterminator='\n')
                writer.writerow(["名前","RT数"])
                writer.writerows(summary_data)
                pass

    #リストを持っていないユーザーのcsvを作成する
    def notHasListUser():
        data = pd.read_csv(properties.csv_path + "リストがない人たち.csv",encoding="utf-8")
        participants.makeUsers()
        for user_id in participants.users_dict:
            data.query('text.str.startswith("RT @' + user_id + ':")', engine='python').to_csv(properties.csv_path + participants.users_dict[user_id] +".csv",index=False )
            dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')
            print("【"+ dt_now_str +"】" + participants.users_dict[user_id] + "のcsv作成完了")

    #CSVをExcelに出力する
    def csvToExcel():
        participants.makeAllUsers()
        for user in participants.all_dict.keys():
            #CSV読み込み
            data = pd.read_csv(properties.csv_path + participants.all_dict[user] + ".csv",encoding="utf-8")
            #最大行数、最大列数の定義
            max_row_num = len(data)
            max_column_num = len(data.columns)
            #出力先のExcel、Sheetを設定
            wb = openpyxl.load_workbook(properties.excel_path)
            sheet = wb[participants.all_dict[user]]
            #現在時刻を取得し、最終更新日時をシートに出力
            dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')
            sheet.cell(row = 1, column = 2, value = "最終更新日時:" + dt_now_str)
            #n行目の処理をする
            for row_num in range(max_row_num):
                #n列目の処理をする
                for column_num in range(max_column_num):
                    #数字ならint型に変える
                    if data[data.columns[column_num]].dtype == 'int64':
                        insert_value = int(data.iloc[[row_num],[column_num]].to_string(index=False, header=False))
                    #数字以外ならそのまま
                    else:
                        insert_value = data.iloc[[row_num],[column_num]].to_string(index=False, header=False)
                    #指定したセルにinsert_valueを出力
                    sheet.cell(row = row_num + 2, column = column_num + 1, value = insert_value)
            #現在時刻を取得し、反映完了ログを出力
            dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')
            print("【"+ dt_now_str +"】" + participants.all_dict[user] + "のデータ反映完了")
            #保存する
            wb.save(properties.excel_path)

    #配信ポイントをCSVに出力する
    def getLivePoint():
        response = urllib.request.urlopen(properties.event_id)
        html = response.read().decode(response.headers.get_content_charset(), errors='ignore')
        parsed_html = BeautifulSoup(html, 'html.parser')
        participants.makeElemsDict()
        dict_num = 1
        mix_data = []
        for elem in participants.elems_dict:
        #ブロックごと変数に格納
            elems = parsed_html.select("#ranking > ul > li:nth-child("+ str(dict_num) + ")")
            #ミクチャID
            mix_id = elems[0].contents[3].attrs['href']
            if mix_id in participants.elems_dict.keys():
                name = participants.elems_dict[mix_id]
                #ミクチャポイント
                mix_pt = elems[0].contents[5].span.text
                mix_data.append([name , mix_pt])
                #CSVに出力
                with open(properties.csv_path + 'mix_point.csv', 'w',newline='',encoding='utf-8') as f:
                    writer = csv.writer(f, lineterminator='\n')
                    writer.writerow(["名前","ポイント"])
                    writer.writerows(mix_data)
                    pass
                #現在時刻を取得し、反映完了ログを出力
                dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')       
            #dict_numを更新
            dict_num += 1
            print("【"+ dt_now_str +"】" + name + "配信ポイントデータ取得完了")

    #配信ポイントCSVをExcelに出力する
    def livePointToExcel():
        participants.makeAllUsers()
        #CSV読み込み
        data = pd.read_csv(properties.csv_path + "mix_point.csv",encoding="utf-8")
        #最大行数、最大列数の定義
        max_row_num = len(data)
        #出力先のExcel、Sheetを設定
        wb = openpyxl.load_workbook(properties.excel_path)
        #行ごとに処理をする
        for row_num in range(max_row_num):
            insert_value = int(data.iloc[[row_num],1].to_string(index=False, header=False).replace(' ', '').replace(',', ''))
            #sheet = wb[data.iloc[[row_num],0].to_string(index=False, header=False)]
            sheet = wb[str(data.iloc[[row_num],0].to_string(index=False, header=False)).replace(' ', '')]
            sheet.cell(row = 1, column = 5, value = insert_value)
        #現在時刻を取得し、反映完了ログを出力
        dt_now_str=datetime.datetime.now().strftime('%Y/%m/%d_%H:%M:%S')
        print("【"+ dt_now_str +"】配信ポイントのデータ反映完了")
        #保存する
        wb.save(properties.excel_path)